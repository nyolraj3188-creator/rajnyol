#!/usr/bin/env python3
"""
Comprehensive backend testing for:
- PART A: Background removal (POST /api/image/remove-bg)
- PART B: Legacy Kruti Dev -> Unicode conversion (pdf-to-word, pdf-to-excel, inspect)
"""
import os
import sys
import io
import requests
from pathlib import Path

# Get backend URL from frontend .env
BACKEND_URL = None
env_file = Path("/app/frontend/.env")
if env_file.exists():
    for line in env_file.read_text().splitlines():
        if line.startswith("REACT_APP_BACKEND_URL="):
            BACKEND_URL = line.split("=", 1)[1].strip()
            break

if not BACKEND_URL:
    print("❌ ERROR: Could not find REACT_APP_BACKEND_URL in /app/frontend/.env")
    sys.exit(1)

print(f"🔗 Testing backend at: {BACKEND_URL}")
print("=" * 80)

# Test counters
total_tests = 0
passed_tests = 0
failed_tests = 0

def test_result(name, passed, details=""):
    global total_tests, passed_tests, failed_tests
    total_tests += 1
    if passed:
        passed_tests += 1
        print(f"✅ {name}")
        if details:
            print(f"   {details}")
    else:
        failed_tests += 1
        print(f"❌ {name}")
        if details:
            print(f"   {details}")
    print()

# ============================================================================
# PART A: Background Removal
# ============================================================================
print("\n" + "=" * 80)
print("PART A: BACKGROUND REMOVAL FIX (POST /api/image/remove-bg)")
print("=" * 80 + "\n")

def test_background_removal():
    """Test POST /api/image/remove-bg with a simple image"""
    from PIL import Image
    
    # Create a simple test image (red square on white background)
    img = Image.new('RGB', (200, 200), color='white')
    # Draw a red square in the center
    for x in range(50, 150):
        for y in range(50, 150):
            img.putpixel((x, y), (255, 0, 0))
    
    # Save to bytes
    img_bytes = io.BytesIO()
    img.save(img_bytes, format='PNG')
    img_bytes.seek(0)
    
    try:
        # POST to remove-bg endpoint
        url = f"{BACKEND_URL}/api/image/remove-bg"
        files = {'file': ('test_image.png', img_bytes, 'image/png')}
        response = requests.post(url, files=files, timeout=60)
        
        # Check HTTP status
        if response.status_code != 200:
            test_result(
                "Background removal - HTTP 200",
                False,
                f"Expected 200, got {response.status_code}: {response.text[:200]}"
            )
            return
        
        test_result("Background removal - HTTP 200", True, f"Status: {response.status_code}")
        
        # Check response is a valid PNG
        content = response.content
        if not content.startswith(b'\x89PNG'):
            test_result(
                "Background removal - Valid PNG signature",
                False,
                f"Response does not start with PNG signature. First 10 bytes: {content[:10]}"
            )
            return
        
        test_result(
            "Background removal - Valid PNG signature",
            True,
            f"Response starts with PNG signature (\\x89PNG)"
        )
        
        # Check size > 100 bytes
        size = len(content)
        if size <= 100:
            test_result(
                "Background removal - Size > 100 bytes",
                False,
                f"Response size is {size} bytes (expected > 100)"
            )
            return
        
        test_result(
            "Background removal - Size > 100 bytes",
            True,
            f"Response size: {size} bytes"
        )
        
        # Verify it's a valid PNG by loading with PIL
        try:
            result_img = Image.open(io.BytesIO(content))
            result_img.verify()
            test_result(
                "Background removal - PIL can load result",
                True,
                f"Image format: {result_img.format}, size: {result_img.size}"
            )
        except Exception as e:
            test_result(
                "Background removal - PIL can load result",
                False,
                f"PIL failed to load: {e}"
            )
        
    except Exception as e:
        test_result("Background removal - Request", False, f"Exception: {e}")

test_background_removal()

# ============================================================================
# PART B: Legacy Kruti Dev -> Unicode Conversion
# ============================================================================
print("\n" + "=" * 80)
print("PART B: LEGACY KRUTI DEV -> UNICODE CONVERSION")
print("=" * 80 + "\n")

def create_kruti_dev_pdf():
    """
    Create a PDF that simulates a real Kruti Dev PDF:
    - Font name contains 'KrutiDev010' (legacy token)
    - Text layer contains ASCII codes (genuine Kruti Dev encoding)
    """
    from fpdf import FPDF
    import re
    
    # Create PDF with custom font name
    pdf = FPDF()
    
    # Register the Lohit Devanagari font under the name 'KrutiDev010'
    # This simulates a legacy font name while using a real Unicode font
    font_path = '/usr/share/fonts/truetype/lohit-devanagari/Lohit-Devanagari.ttf'
    if not Path(font_path).exists():
        raise FileNotFoundError(f"Font not found: {font_path}")
    
    pdf.add_font('KrutiDev010', '', font_path, uni=True)
    pdf.add_page()
    pdf.set_font('KrutiDev010', size=20)
    
    # Write EXACT ASCII lines (genuine Kruti Dev encoding)
    # These should convert to real Devanagari
    kruti_lines = [
        'pfj= izek.k i=',
        'izekf.kr fd;k tkrk gS fd Jh@dqekjh@Jherh',
        'O;fäxr :i ls ekg@o"kksZa ls tkurk@tkurh gwi rFkk tgki rd esjl',
        'budk uSfrd pfj= mmke gSA'
    ]
    
    for line in kruti_lines:
        pdf.cell(0, 10, txt=line, ln=True)
    
    # Save to bytes
    pdf_bytes = pdf.output()
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode('latin-1')
    
    # Manually patch the PDF to inject the legacy font name
    # Replace the auto-generated font name with 'KrutiDev010'
    pdf_str = pdf_bytes.decode('latin-1', errors='ignore')
    # Find and replace the font name in the PDF structure
    # fpdf2 generates names like 'MPDFAA+LohitDevanagari'
    pdf_str = re.sub(r'/BaseFont\s*/[A-Z]+\+LohitDevanagari', '/BaseFont /KrutiDev010', pdf_str)
    pdf_str = re.sub(r'/FontName\s*/[A-Z]+\+LohitDevanagari', '/FontName /KrutiDev010', pdf_str)
    pdf_bytes = pdf_str.encode('latin-1', errors='ignore')
    
    return pdf_bytes

def has_devanagari(text):
    """Check if text contains Devanagari Unicode characters (U+0900-U+097F)"""
    if not text:
        return False
    return any('\u0900' <= c <= '\u097f' for c in text)

def count_devanagari(text):
    """Count Devanagari Unicode characters"""
    if not text:
        return 0
    return sum(1 for c in text if '\u0900' <= c <= '\u097f')

def extract_docx_text(docx_bytes):
    """Extract text from .docx bytes"""
    from docx import Document
    doc = Document(io.BytesIO(docx_bytes))
    return '\n'.join(p.text for p in doc.paragraphs)

def extract_xlsx_text(xlsx_bytes):
    """Extract all cell text from .xlsx bytes"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    text_parts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    text_parts.append(str(cell.value))
    return '\n'.join(text_parts)

# Test 1: PDF to Word with legacy Kruti Dev
print("Test 1: POST /api/pdf/pdf-to-word with legacy Kruti Dev PDF")
print("-" * 80)

try:
    pdf_bytes = create_kruti_dev_pdf()
    print(f"✓ Created Kruti Dev PDF ({len(pdf_bytes)} bytes)")
    
    url = f"{BACKEND_URL}/api/pdf/pdf-to-word"
    files = {'file': ('kruti_test.pdf', io.BytesIO(pdf_bytes), 'application/pdf')}
    response = requests.post(url, files=files, timeout=120)
    
    if response.status_code != 200:
        test_result(
            "Kruti Dev PDF to Word - HTTP 200",
            False,
            f"Expected 200, got {response.status_code}: {response.text[:300]}"
        )
    else:
        test_result("Kruti Dev PDF to Word - HTTP 200", True)
        
        # Check it's a valid .docx
        docx_bytes = response.content
        if len(docx_bytes) < 1000:
            test_result(
                "Kruti Dev PDF to Word - Valid non-empty .docx",
                False,
                f"Response too small: {len(docx_bytes)} bytes"
            )
        else:
            test_result(
                "Kruti Dev PDF to Word - Valid non-empty .docx",
                True,
                f"Size: {len(docx_bytes)} bytes"
            )
            
            # Extract text
            try:
                text = extract_docx_text(docx_bytes)
                text_len = len(text.strip())
                
                test_result(
                    "Kruti Dev PDF to Word - Non-empty text",
                    text_len > 0,
                    f"Extracted {text_len} characters"
                )
                
                # Check for Devanagari Unicode
                deva_count = count_devanagari(text)
                has_deva = has_devanagari(text)
                
                if not has_deva:
                    test_result(
                        "Kruti Dev PDF to Word - Contains Devanagari (U+0900-U+097F)",
                        False,
                        f"No Devanagari found. Text preview: {text[:200]}"
                    )
                else:
                    test_result(
                        "Kruti Dev PDF to Word - Contains Devanagari (U+0900-U+097F)",
                        True,
                        f"Found {deva_count} Devanagari characters"
                    )
                    print(f"   Text preview: {text[:200]}")
                
                # Check that ASCII gibberish is NOT present
                ascii_gibberish = ['pfj=', 'izekf.kr', 'O;fäxr', 'budk']
                found_gibberish = [g for g in ascii_gibberish if g in text]
                
                if found_gibberish:
                    test_result(
                        "Kruti Dev PDF to Word - No ASCII gibberish",
                        False,
                        f"Found ASCII gibberish: {found_gibberish}"
                    )
                else:
                    test_result(
                        "Kruti Dev PDF to Word - No ASCII gibberish",
                        True,
                        "No ASCII tokens like 'pfj=' or 'izekf.kr' found"
                    )
                
            except Exception as e:
                test_result(
                    "Kruti Dev PDF to Word - Extract text",
                    False,
                    f"Failed to extract text: {e}"
                )
                
except Exception as e:
    test_result("Kruti Dev PDF to Word - Create/POST", False, f"Exception: {e}")

# Test 2: PDF to Excel with legacy Kruti Dev
print("\nTest 2: POST /api/pdf/pdf-to-excel with legacy Kruti Dev PDF")
print("-" * 80)

try:
    pdf_bytes = create_kruti_dev_pdf()
    
    url = f"{BACKEND_URL}/api/pdf/pdf-to-excel"
    files = {'file': ('kruti_test.pdf', io.BytesIO(pdf_bytes), 'application/pdf')}
    response = requests.post(url, files=files, timeout=120)
    
    if response.status_code != 200:
        test_result(
            "Kruti Dev PDF to Excel - HTTP 200",
            False,
            f"Expected 200, got {response.status_code}: {response.text[:300]}"
        )
    else:
        test_result("Kruti Dev PDF to Excel - HTTP 200", True)
        
        # Check it's a valid .xlsx
        xlsx_bytes = response.content
        if len(xlsx_bytes) < 1000:
            test_result(
                "Kruti Dev PDF to Excel - Valid non-empty .xlsx",
                False,
                f"Response too small: {len(xlsx_bytes)} bytes"
            )
        else:
            test_result(
                "Kruti Dev PDF to Excel - Valid non-empty .xlsx",
                True,
                f"Size: {len(xlsx_bytes)} bytes"
            )
            
            # Extract text from cells
            try:
                text = extract_xlsx_text(xlsx_bytes)
                
                # Check for Devanagari Unicode
                deva_count = count_devanagari(text)
                has_deva = has_devanagari(text)
                
                if not has_deva:
                    test_result(
                        "Kruti Dev PDF to Excel - Contains Devanagari (U+0900-U+097F)",
                        False,
                        f"No Devanagari found. Text preview: {text[:200]}"
                    )
                else:
                    test_result(
                        "Kruti Dev PDF to Excel - Contains Devanagari (U+0900-U+097F)",
                        True,
                        f"Found {deva_count} Devanagari characters in cells"
                    )
                    print(f"   Text preview: {text[:200]}")
                
            except Exception as e:
                test_result(
                    "Kruti Dev PDF to Excel - Extract text",
                    False,
                    f"Failed to extract text: {e}"
                )
                
except Exception as e:
    test_result("Kruti Dev PDF to Excel - Create/POST", False, f"Exception: {e}")

# Test 3: PDF inspect endpoint
print("\nTest 3: POST /api/pdf/inspect with legacy Kruti Dev PDF")
print("-" * 80)

try:
    pdf_bytes = create_kruti_dev_pdf()
    
    url = f"{BACKEND_URL}/api/pdf/inspect"
    files = {'file': ('kruti_test.pdf', io.BytesIO(pdf_bytes), 'application/pdf')}
    response = requests.post(url, files=files, timeout=60)
    
    if response.status_code != 200:
        test_result(
            "Kruti Dev PDF inspect - HTTP 200",
            False,
            f"Expected 200, got {response.status_code}: {response.text[:300]}"
        )
    else:
        test_result("Kruti Dev PDF inspect - HTTP 200", True)
        
        data = response.json()
        legacy_hindi = data.get('legacy_hindi', False)
        
        if not legacy_hindi:
            test_result(
                "Kruti Dev PDF inspect - legacy_hindi == true",
                False,
                f"Expected legacy_hindi=true, got {legacy_hindi}. Response: {data}"
            )
        else:
            test_result(
                "Kruti Dev PDF inspect - legacy_hindi == true",
                True,
                f"Response: {data}"
            )
            
except Exception as e:
    test_result("Kruti Dev PDF inspect - POST", False, f"Exception: {e}")

# Test 4: Regression - English PDF
print("\nTest 4: REGRESSION - Plain English PDF")
print("-" * 80)

def create_english_pdf():
    """Create a simple English text PDF"""
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', size=12)
    pdf.cell(0, 10, txt='This is a test document in English.', ln=True)
    pdf.cell(0, 10, txt='It should convert to Word without issues.', ln=True)
    pdf.cell(0, 10, txt='No Hindi or Devanagari text here.', ln=True)
    pdf_bytes = pdf.output()
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode('latin-1')
    return pdf_bytes

try:
    pdf_bytes = create_english_pdf()
    
    # Test inspect
    url = f"{BACKEND_URL}/api/pdf/inspect"
    files = {'file': ('english_test.pdf', io.BytesIO(pdf_bytes), 'application/pdf')}
    response = requests.post(url, files=files, timeout=60)
    
    if response.status_code == 200:
        data = response.json()
        legacy_hindi = data.get('legacy_hindi', False)
        
        if legacy_hindi:
            test_result(
                "English PDF inspect - legacy_hindi == false",
                False,
                f"Expected legacy_hindi=false, got {legacy_hindi}"
            )
        else:
            test_result(
                "English PDF inspect - legacy_hindi == false",
                True,
                f"Response: {data}"
            )
    
    # Test pdf-to-word
    url = f"{BACKEND_URL}/api/pdf/pdf-to-word"
    files = {'file': ('english_test.pdf', io.BytesIO(pdf_bytes), 'application/pdf')}
    response = requests.post(url, files=files, timeout=120)
    
    if response.status_code != 200:
        test_result(
            "English PDF to Word - HTTP 200",
            False,
            f"Expected 200, got {response.status_code}"
        )
    else:
        test_result("English PDF to Word - HTTP 200", True)
        
        docx_bytes = response.content
        text = extract_docx_text(docx_bytes)
        
        if 'English' not in text or len(text.strip()) < 10:
            test_result(
                "English PDF to Word - Valid English text",
                False,
                f"Text: {text[:200]}"
            )
        else:
            test_result(
                "English PDF to Word - Valid English text",
                True,
                f"Extracted {len(text.strip())} characters"
            )
            
except Exception as e:
    test_result("English PDF regression - POST", False, f"Exception: {e}")

# Test 5: Health check
print("\nTest 5: GET /api/pdf/health")
print("-" * 80)

try:
    url = f"{BACKEND_URL}/api/pdf/health"
    response = requests.get(url, timeout=30)
    
    if response.status_code != 200:
        test_result(
            "PDF health check - HTTP 200",
            False,
            f"Expected 200, got {response.status_code}"
        )
    else:
        test_result("PDF health check - HTTP 200", True)
        
        data = response.json()
        tools = data.get('tools', {})
        
        required_tools = ['soffice', 'gs', 'qpdf', 'tesseract', 'pdftoppm', 'ocrmypdf']
        all_available = all(tools.get(t, False) for t in required_tools)
        
        if not all_available:
            missing = [t for t in required_tools if not tools.get(t, False)]
            test_result(
                "PDF health check - All tools available",
                False,
                f"Missing tools: {missing}. Tools: {tools}"
            )
        else:
            test_result(
                "PDF health check - All tools available",
                True,
                f"All required tools available: {tools}"
            )
            
except Exception as e:
    test_result("PDF health check - GET", False, f"Exception: {e}")

# ============================================================================
# SUMMARY
# ============================================================================
print("\n" + "=" * 80)
print("TEST SUMMARY")
print("=" * 80)
print(f"Total tests: {total_tests}")
print(f"✅ Passed: {passed_tests}")
print(f"❌ Failed: {failed_tests}")
print(f"Success rate: {(passed_tests/total_tests*100):.1f}%")
print("=" * 80)

if failed_tests == 0:
    print("\n🎉 ALL TESTS PASSED!")
    sys.exit(0)
else:
    print(f"\n⚠️  {failed_tests} TEST(S) FAILED")
    sys.exit(1)
